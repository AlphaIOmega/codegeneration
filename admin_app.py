# File: admin_app.py
from flask import Flask, render_template, request, redirect, url_for, jsonify, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import bcrypt
import openpyxl
import os
import csv
from threading import Lock
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Flask app
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "default_secret_key")

# Flask-Login setup
login_manager = LoginManager(app)
login_manager.login_view = "login"

# Paths to files
EXCEL_FILE = "codes.xlsx"
LOG_FILE = "user_logs.csv"

# Lock for thread safety
lock = Lock()

# Admin credentials
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD_HASH = os.getenv("ADMIN_PASSWORD_HASH", bcrypt.hashpw(b"password", bcrypt.gensalt()))

# Flask-Login user class
class User(UserMixin):
    def __init__(self, id):
        self.id = id

# User loader for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    if user_id == ADMIN_USERNAME:
        return User(user_id)
    return None

# Helper: Load user logs
def load_user_logs():
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Username", "Code", "Timestamp"])
    with open(LOG_FILE, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        return list(reader)

# Helper: Write user logs
def write_user_logs(logs):
    with open(LOG_FILE, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Username", "Code", "Timestamp"])
        writer.writerows([[log["Username"], log["Code"], log["Timestamp"]] for log in logs])

# Helper: Load unused codes
def load_unused_codes():
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    unused_codes = []
    for row in sheet.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value and not cell.font.strike:
            unused_codes.append(cell.value)
    return unused_codes

# Helper: Mark a code as unused
def mark_code_as_unused(code):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value == code:
            cell.font = openpyxl.styles.Font(strike=False)
            wb.save(EXCEL_FILE)
            return True
    return False

# Routes
@app.route("/")
@login_required
def admin_dashboard():
    return render_template("admin.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"].encode("utf-8")
        if username == ADMIN_USERNAME and bcrypt.checkpw(password, ADMIN_PASSWORD_HASH):
            login_user(User(username))
            return redirect(url_for("admin_dashboard"))
        else:
            flash("Invalid username or password")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

@app.route("/api/logs", methods=["GET", "DELETE"])
@login_required
def manage_logs():
    if request.method == "GET":
        logs = load_user_logs()
        return jsonify(logs)
    elif request.method == "DELETE":
        data = request.json
        username = data.get("username")
        code = data.get("code")
        with lock:
            logs = load_user_logs()
            logs = [log for log in logs if not (log["Username"] == username and log["Code"] == code)]
            write_user_logs(logs)
        return jsonify({"success": True, "message": "Log entry deleted."})

@app.route("/api/unused-codes", methods=["GET", "POST"])
@login_required
def manage_unused_codes():
    if request.method == "GET":
        unused_codes = load_unused_codes()
        return jsonify(unused_codes)
    elif request.method == "POST":
        data = request.json
        code = data.get("code")
        with lock:
            if mark_code_as_unused(code):
                return jsonify({"success": True, "message": f"Code {code} marked as unused."})
            else:
                return jsonify({"success": False, "message": f"Code {code} not found."})

if __name__ == "__main__":
    app.run(debug=True)
